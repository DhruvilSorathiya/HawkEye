import csv
import html
import json
import logging
import multiprocessing as mp
import os
import random
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

import cherrypy
from cherrypy import _cperror

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl

import secure

from helib import HawkEye

from hescan import startHawkEyeScanner

from hawkeye import HawkEyeDb
from hawkeye import HawkEyeHelpers
from hawkeye import __version__
from hawkeye.logger import logListenerSetup, logWorkerSetup
from hawkeye.auth import HawkEyeAuth

mp.set_start_method("spawn", force=True)
print("Looking for template in:", os.path.abspath('hawkeye/templates/admin_dashboard.tmpl'))
print("Exists:", os.path.isfile('hawkeye/templates/admin_dashboard.tmpl'))

class HawkEyeWebUi:
    """HawkEye web interface."""

    lookup = TemplateLookup(
        directories=[os.path.join(os.path.dirname(os.path.abspath(__file__)), 'hawkeye', 'templates')]
    )
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'HawkEyeWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): HawkEye config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = HawkEyeDb(self.defaultConfig, init=True)
        he = HawkEye(self.defaultConfig)
        self.config = he.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"hawkeye.{__name__}")
        
        # Initialize authentication
        self.auth = HawkEyeAuth(dbh)

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:", "https://unpkg.com")
            .style_src("'self'", "'unsafe-inline'", "https://unpkg.com")
            .base_uri("'self'")
            .connect_src("'self'", "data:", "https://*.tile.openstreetmap.org")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:", "https://*.tile.openstreetmap.org")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'HawkEyeWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'HawkEyeWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'HawkEyeWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = self.lookup.get_template('error.tmpl')
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'HawkEyeWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'HawkEyeWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = self.lookup.get_template('error.tmpl')
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'HawkEyeWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'HawkEyeWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = HawkEyeDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'HawkEyeWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # AUTHENTICATION PAGES
    #
    
    @cherrypy.expose
    def login(self, username=None, password=None) -> str:
        """Show login page or process admin login."""

        from mako.template import Template
        import cherrypy
        import traceback

        try:
            # If the form is submitted
            if username is not None and password is not None:
                # Simple admin check
                if username == "admin" and password == "admin":
                    # Redirect to admin dashboard
                    raise cherrypy.HTTPRedirect("/admin_dashboard")
                else:
                    # Invalid login, render error page
                    templ = self.lookup.get_template('error.tmpl')
                    return templ.render(error="Invalid username or password")

            # If form not submitted, show login selection page
            templ = self.lookup.get_template('login_select.tmpl')
            return templ.render(
                docroot=(getattr(self, "docroot", "") or ""),
                version=getattr(__import__("hawkeye"), "__version__", "1.0")
            )

        except Exception as e:
            print("[ERROR] Exception in login():", e)
            traceback.print_exc()
            return f"<h2>Error in login()</h2><pre>{str(e)}</pre>"
    
    @cherrypy.expose
    def default(self) -> str:
        """Default route - redirect to login if not authenticated.
        
        Returns:
            str: Redirect to appropriate page
        """
        # Check if user is authenticated
        if self.auth.require_auth():
            # User is authenticated, redirect to scans page
            raise cherrypy.HTTPRedirect(f"{self.docroot}/")
        else:
            # User is not authenticated, redirect to login
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
    
    @cherrypy.expose
    def login_user(self, username: str = None, password: str = None) -> str:
        """User login page and handler.
        
        Args:
            username (str): username
            password (str): password
            
        Returns:
            str: Login page HTML or redirect
        """
        error = None
        success = None
        
        if cherrypy.request.method == 'POST':
            if not username or not password:
                error = "Please enter both username and password"
            else:
                ip_address = self.auth.get_client_ip()
                user_agent = self.auth.get_user_agent()
                
                if self.auth.login_user(username, password, ip_address, user_agent):
                    # Log activity
                    dbh = HawkEyeDb(self.config)
                    user_info = self.auth.get_current_user()
                    if user_info:
                        dbh.logUserActivity(
                            user_id=user_info['id'],
                            activity_type='login',
                            activity_description='User logged in successfully',
                            ip_address=ip_address
                        )
                    raise cherrypy.HTTPRedirect(f"{self.docroot}/")
                else:
                    error = "Invalid username or password"
        
        templ = self.lookup.get_template('login_user.tmpl')
        return templ.render(docroot=self.docroot, version=__version__, error=error, success=success)
    
    @cherrypy.expose
    def login_admin(self, username: str = None, password: str = None) -> str:
        """Admin login page and handler.
        
        Args:
            username (str): username
            password (str): password
            
        Returns:
            str: Login page HTML or redirect
        """
        error = None
        success = None
        
        if cherrypy.request.method == 'POST':
            if not username or not password:
                error = "Please enter both username and password"
            else:
                ip_address = self.auth.get_client_ip()
                user_agent = self.auth.get_user_agent()
                
                if self.auth.login_admin(username, password, ip_address, user_agent):
                    raise cherrypy.HTTPRedirect(f"{self.docroot}/admin")
                else:
                    error = "Invalid admin username or password"
        
        templ = self.lookup.get_template('login_admin.tmpl')
        return templ.render(docroot=self.docroot, version=__version__, error=error, success=success)
    
    @cherrypy.expose
    def logout(self) -> str:
        """Logout handler.
        
        Returns:
            str: Redirect to login page
        """
        self.auth.logout()
        raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
    
    @cherrypy.expose
    def forgot_password(self, email: str = None, user_type: str = 'user') -> str:
        """Forgot password page and handler.
        
        Args:
            email (str): email address
            user_type (str): 'user' or 'admin'
            
        Returns:
            str: Forgot password page HTML
        """
        error = None
        success = None
        
        if cherrypy.request.method == 'POST':
            if not email:
                error = "Please enter your email address"
            else:
                dbh = HawkEyeDb(self.config)
                
                # Check if user/admin exists with this email
                if user_type == 'admin':
                    account = dbh.getAdminByEmail(email)
                else:
                    account = dbh.getUserByEmail(email)
                
                if account:
                    # Generate reset token
                    if user_type == 'admin':
                        token = dbh.createPasswordResetToken(admin_id=account['id'])
                    else:
                        token = dbh.createPasswordResetToken(user_id=account['id'])
                    
                    # Create reset link
                    reset_link = f"{cherrypy.request.base}{self.docroot}/reset_password?token={token}"
                    
                    # Try to send email if configured
                    email_sent = False
                    email_config = self.config.get('email', {})
                    
                    if email_config.get('smtp_user') and email_config.get('smtp_password'):
                        try:
                            from hawkeye.email_helper import EmailHelper
                            
                            email_helper = EmailHelper(
                                smtp_host=email_config.get('smtp_host', 'smtp.gmail.com'),
                                smtp_port=email_config.get('smtp_port', 587),
                                smtp_user=email_config.get('smtp_user'),
                                smtp_password=email_config.get('smtp_password'),
                                from_email=email_config.get('from_email', email_config.get('smtp_user'))
                            )
                            
                            email_success, email_message = email_helper.send_password_reset_email(
                                to_email=email,
                                username=account['username'],
                                reset_link=reset_link
                            )
                            
                            if email_success:
                                email_sent = True
                                success = "Password reset link has been sent to your email. Please check your inbox."
                            else:
                                error = f"Failed to send email: {email_message}. Please contact administrator."
                                self.log.error(f"Email send failed: {email_message}")
                        except Exception as e:
                            self.log.error(f"Email error: {str(e)}")
                            error = "Failed to send email. Please contact administrator."
                    
                    # If email not sent, show the link (development mode)
                    if not email_sent:
                        success = f"<strong>Development Mode:</strong> Email not configured. Copy this link to reset your password:<br><br>" \
                                 f"<div style='background: #f5f5f5; padding: 10px; border-radius: 5px; word-break: break-all; margin: 10px 0;'>" \
                                 f"<a href='{reset_link}' target='_blank'>{reset_link}</a></div>" \
                                 f"<small>This link expires in 1 hour.</small>"
                    
                    # Log the reset request
                    self.log.info(f"Password reset requested for {user_type}: {email}")
                else:
                    # Don't reveal if email exists or not for security
                    success = "If an account with this email exists, a password reset link has been sent."
        
        templ = self.lookup.get_template('forgot_password.tmpl')
        return templ.render(docroot=self.docroot, version=__version__, error=error, success=success, user_type=user_type)
    
    @cherrypy.expose
    def reset_password(self, token: str = None, new_password: str = None, confirm_password: str = None) -> str:
        """Reset password page and handler.
        
        Args:
            token (str): password reset token
            new_password (str): new password
            confirm_password (str): confirm new password
            
        Returns:
            str: Reset password page HTML
        """
        error = None
        success = None
        
        if not token:
            error = "Invalid or missing reset token"
            templ = self.lookup.get_template('reset_password.tmpl')
            return templ.render(docroot=self.docroot, version=__version__, error=error, success=success, token=None)
        
        dbh = HawkEyeDb(self.config)
        token_info = dbh.getPasswordResetToken(token)
        
        if not token_info:
            error = "Invalid or expired reset token"
            templ = self.lookup.get_template('reset_password.tmpl')
            return templ.render(docroot=self.docroot, version=__version__, error=error, success=success, token=None)
        
        if cherrypy.request.method == 'POST':
            if not new_password or not confirm_password:
                error = "Please enter and confirm your new password"
            elif new_password != confirm_password:
                error = "Passwords do not match"
            elif len(new_password) < 6:
                error = "Password must be at least 6 characters long"
            else:
                # Reset the password
                if token_info['user_id']:
                    dbh.resetUserPassword(token_info['user_id'], new_password)
                elif token_info['admin_id']:
                    dbh.resetAdminPassword(token_info['admin_id'], new_password)
                
                # Mark token as used
                dbh.markPasswordResetTokenUsed(token)
                
                success = "Password reset successfully! You can now login with your new password."
                self.log.info(f"Password reset completed for token: {token}")
        
        templ = self.lookup.get_template('reset_password.tmpl')
        return templ.render(docroot=self.docroot, version=__version__, error=error, success=success, token=token)
    
    @cherrypy.expose
    def admin(self) -> str:
        """Admin dashboard.
        
        Returns:
            str: Admin dashboard HTML
        """
        if not self.auth.require_auth('admin'):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login/admin")
        
        dbh = HawkEyeDb(self.config)
        
        # Get admin statistics
        admins = dbh.getAllAdmins()
        admins_fmt = []
        for a in admins:
            created_val = a[4] if len(a) > 4 else None
            last_login_val = a[5] if len(a) > 5 else None
            admins_fmt.append([
                a[0],            # id
                a[1],            # username
                a[2],            # email
                a[3] or 'N/A',   # mobile
                created_val,     # created (epoch or None)
                last_login_val,  # last login (epoch or None)
                a[6] if len(a) > 6 else False  # is_active
            ])
        total_admins = len(admins)
        
        # Get user statistics
        users = dbh.getAllUsers()
        users_fmt = []
        for u in users:
            created_val = u[4] if len(u) > 4 else None
            last_login_val = u[5] if len(u) > 5 else None
            users_fmt.append([
                u[0],            # id
                u[1],            # username
                u[2],            # email
                u[3] or 'N/A',   # mobile
                created_val,     # created (epoch or None)
                last_login_val,  # last login (epoch or None)
                u[6] if len(u) > 6 else False  # is_active
            ])
        total_users = len(users)
        
        # Get scan statistics (admin panel shows all scans)
        scan_list = dbh.scanInstanceList(is_admin_panel=True)
        total_scans = len(scan_list)
        active_scans = len([s for s in scan_list if s[6] in ['RUNNING', 'STARTING', 'STARTED']])
        
        # Get today's logins
        today_logins = dbh.getTodayLogins()
        
        # Get current user info
        user_info = self.auth.get_current_user()
        
        templ = self.lookup.get_template('admin_dashboard.tmpl')
        return templ.render(
            docroot=self.docroot, 
            version=__version__,
            admins=admins_fmt,
            total_admins=total_admins,
            users=users_fmt,
            total_users=total_users,
            total_scans=total_scans,
            active_scans=active_scans,
            today_logins=today_logins,
            pageid='ADMIN',
            time=time,
            user_info=user_info
        )
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_create_user(self, username: str, password: str, email: str, mobile: str = None) -> dict:
        """Create a new user (admin only).
        
        Args:
            username (str): username
            password (str): password
            email (str): email
            mobile (str): mobile number
            
        Returns:
            dict: JSON response
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        if not username or not password or not email:
            return {'success': False, 'message': 'Username, password, and email are required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            current_admin = self.auth.get_current_user()
            
            if dbh.createUser(username, password, email, mobile, current_admin['id']):
                return {'success': True, 'message': 'User created successfully'}
            else:
                return {'success': False, 'message': 'Failed to create user'}
        except Exception as e:
            return {'success': False, 'message': f'Error creating user: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_create_admin(self, username: str, password: str, email: str, mobile: str = None, 
                          auth_username: str = None, auth_password: str = None, 
                          confirm_password: str = None) -> dict:
        """Create a new admin (admin only).
        
        Args:
            username (str): username for new admin
            password (str): password for new admin
            email (str): email for new admin
            mobile (str): mobile number (optional)
            auth_username (str): authorizing admin username
            auth_password (str): authorizing admin password
            confirm_password (str): password confirmation
            
        Returns:
            dict: JSON response
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        if not username or not password or not email:
            return {'success': False, 'message': 'Username, password, and email are required'}
        
        if not auth_username or not auth_password:
            return {'success': False, 'message': 'Authorizing admin credentials are required'}
        
        if confirm_password and password != confirm_password:
            return {'success': False, 'message': 'Passwords do not match'}
        
        try:
            dbh = HawkEyeDb(self.config)
            
            # Verify authorizing admin credentials
            ip_address = self.auth.get_client_ip()
            user_agent = self.auth.get_user_agent()
            if not self.auth.login_admin(auth_username, auth_password, ip_address, user_agent, verify_only=True):
                return {'success': False, 'message': 'Invalid authorizing admin credentials'}
            
            current_admin = self.auth.get_current_user()
            
            if dbh.createAdmin(username, password, email, mobile, current_admin['id']):
                return {'success': True, 'message': 'Admin created successfully'}
            else:
                return {'success': False, 'message': 'Failed to create admin'}
        except Exception as e:
            return {'success': False, 'message': f'Error creating admin: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_user_activity(self, user_id: str) -> dict:
        """Get user activity (admin only).
        
        Args:
            user_id (str): user ID
            
        Returns:
            dict: JSON response with user activity
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            activities = dbh.getUserActivity(int(user_id))
            
            return {
                'success': True,
                'activities': [
                    {
                        'activity_type': activity[0],
                        'activity_description': activity[1],
                        'scan_id': activity[2],
                        'created_at': activity[3],
                        'ip_address': activity[4]
                    }
                    for activity in activities
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching user activity: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_toggle_user_status(self, user_id: str, action: str) -> dict:
        """Toggle user status (admin only).
        
        Args:
            user_id (str): user ID
            action (str): 'activate' or 'deactivate'
            
        Returns:
            dict: JSON response
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            is_active = 1 if action == 'activate' else 0
            
            with dbh.dbhLock:
                dbh.dbh.execute("UPDATE tbl_users SET is_active = ? WHERE id = ?", (is_active, int(user_id)))
                dbh.conn.commit()
            
            return {'success': True, 'message': f'User {action}d successfully'}
        except Exception as e:
            return {'success': False, 'message': f'Error updating user status: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_admins(self) -> dict:
        """Get all admins list (admin only).
        
        Returns:
            dict: JSON response with admin list
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            admins = dbh.getAllAdmins()
            
            return {
                'success': True,
                'admins': [
                    {
                        'id': admin[0],
                        'username': admin[1],
                        'email': admin[2],
                        'mobile': admin[3] if admin[3] else 'N/A',
                        'created_at': admin[4],
                        'last_login': admin[5],
                        'is_active': admin[6] if len(admin) > 6 else True
                    }
                    for admin in admins
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching admins: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_users(self) -> dict:
        """Get all users list (admin only).
        
        Returns:
            dict: JSON response with user list
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            users = dbh.getAllUsers()
            
            return {
                'success': True,
                'users': [
                    {
                        'id': user[0],
                        'username': user[1],
                        'email': user[2],
                        'mobile': user[3] if user[3] else 'N/A',
                        'created_at': user[4],
                        'last_login': user[5],
                        'is_active': user[6] if len(user) > 6 else True
                    }
                    for user in users
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching users: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_today_logins(self) -> dict:
        """Get today's login details (admin only).
        
        Returns:
            dict: JSON response with today's login details
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            logins = dbh.getTodayLoginDetails()
            
            return {
                'success': True,
                'logins': [
                    {
                        'username': login[0],
                        'login_time': login[1],
                        'ip_address': login[2] if login[2] else 'N/A',
                        'account_type': login[3]
                    }
                    for login in logins
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching today\'s logins: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_all_scans(self) -> dict:
        """Get all scans with user information (admin only).
        
        Returns:
            dict: JSON response with all scans
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            scans = dbh.getAllScansWithOwners()
            
            return {
                'success': True,
                'scans': scans
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching scans: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_active_scans(self) -> dict:
        """Get active/running scans (admin only).
        
        Returns:
            dict: JSON response with active scans
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            scans = dbh.getActiveScansWithOwners()
            
            return {
                'success': True,
                'scans': scans
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching active scans: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_recent_activity(self) -> dict:
        """Get recent system activity (admin only).
        
        Returns:
            dict: JSON response with recent activity
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            
            # Get recent user activities (last 10)
            with dbh.dbhLock:
                dbh.dbh.execute("""
                    SELECT ua.activity_type, ua.activity_description, ua.scan_id, ua.created_at, ua.ip_address, u.username
                    FROM tbl_user_activity ua
                    JOIN tbl_users u ON ua.user_id = u.id
                    ORDER BY ua.created_at DESC
                    LIMIT 10
                """)
                activities = dbh.dbh.fetchall()
            
            return {
                'success': True,
                'activities': [
                    {
                        'activity_type': activity[0],
                        'activity_description': f"{activity[5]}: {activity[1]}",
                        'scan_id': activity[2],
                        'created_at': activity[3],
                        'ip_address': activity[4]
                    }
                    for activity in activities
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching recent activity: {str(e)}'}
    
    @cherrypy.expose
    def admin_export_users(self) -> str:
        """Export user data (admin only).
        
        Returns:
            str: CSV export of user data
        """
        if not self.auth.require_auth('admin'):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login/admin")
        
        try:
            dbh = HawkEyeDb(self.config)
            users = dbh.getAllUsers()
            
            csv_content = "Username,Email,Mobile,Created,Last Login,Status\n"
            for user in users:
                created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(user[4]))
                last_login = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(user[5])) if user[5] else "Never"
                status = "Active" if user[6] else "Inactive"
                csv_content += f"{user[1]},{user[2]},{user[3] or ''},{created},{last_login},{status}\n"
            
            cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="hawkeye_users.csv"'
            cherrypy.response.headers['Content-Type'] = 'text/csv'
            return csv_content
        except Exception as e:
            return f"Error exporting users: {str(e)}"
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_system_settings(self) -> dict:
        """Get all system settings (admin only).
        
        Returns:
            dict: JSON response with system settings
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            settings = dbh.getSystemSettings()
            
            return {
                'success': True,
                'settings': [
                    {
                        'id': setting[0],
                        'key': setting[1],
                        'value': setting[2],
                        'type': setting[3],
                        'description': setting[4],
                        'updated_at': setting[5],
                        'updated_by': setting[6]
                    }
                    for setting in settings
                ]
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching system settings: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_update_system_setting(self, setting_key: str, setting_value: str) -> dict:
        """Update a system setting (admin only).
        
        Args:
            setting_key (str): setting key
            setting_value (str): new value
            
        Returns:
            dict: JSON response
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            current_admin = self.auth.get_current_user()
            
            if dbh.updateSystemSetting(setting_key, setting_value, current_admin['id']):
                return {'success': True, 'message': 'Setting updated successfully'}
            else:
                return {'success': False, 'message': 'Failed to update setting'}
        except Exception as e:
            return {'success': False, 'message': f'Error updating setting: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_system_logs(self, limit: int = 100, offset: int = 0, log_level: str = None, 
                             log_category: str = None, start_date: int = None, end_date: int = None) -> dict:
        """Get system logs with filtering (admin only).
        
        Args:
            limit (int): number of logs to return
            offset (int): offset for pagination
            log_level (str): filter by log level
            log_category (str): filter by category
            start_date (int): start timestamp
            end_date (int): end timestamp
            
        Returns:
            dict: JSON response with system logs
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            
            # Convert string parameters to appropriate types
            limit = int(limit) if limit else 100
            offset = int(offset) if offset else 0
            start_date = int(start_date) if start_date else None
            end_date = int(end_date) if end_date else None
            
            result = dbh.getSystemLogs(limit, offset, log_level, log_category, start_date, end_date)
            
            return {
                'success': True,
                'logs': [
                    {
                        'id': log[0],
                        'level': log[1],
                        'category': log[2],
                        'message': log[3],
                        'user_id': log[4],
                        'admin_id': log[5],
                        'ip_address': log[6],
                        'user_agent': log[7],
                        'additional_data': log[8],
                        'created_at': log[9],
                        'user_username': log[10],
                        'admin_username': log[11]
                    }
                    for log in result['logs']
                ],
                'total': result['total']
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching system logs: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def admin_get_system_stats(self) -> dict:
        """Get system statistics (admin only).
        
        Returns:
            dict: JSON response with system stats
        """
        if not self.auth.require_auth('admin'):
            return {'success': False, 'message': 'Admin authentication required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            stats = dbh.getSystemStats()
            
            return {
                'success': True,
                'stats': stats
            }
        except Exception as e:
            return {'success': False, 'message': f'Error fetching system stats: {str(e)}'}
    
    @cherrypy.expose
    def profile(self) -> str:
        """User/Admin profile page.
        
        Returns:
            str: profile page HTML
        """
        # Check if user is logged in
        user_info = self.auth.get_current_user()
        if not user_info:
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
        
        dbh = HawkEyeDb(self.config)
        
        # Get profile data
        profile_data = {
            'username': user_info['username'],
            'email': user_info['email'],
            'mobile': user_info.get('mobile', 'N/A'),
            'type': user_info['type'],
            'is_active': True
        }
        
        # Get additional user/admin data
        if user_info['type'] == 'admin':
            with dbh.dbhLock:
                dbh.dbh.execute("""
                    SELECT created_at, last_login, is_active 
                    FROM tbl_admins WHERE id = ?
                """, (user_info['id'],))
                result = dbh.dbh.fetchone()
                if result:
                    profile_data['created_at'] = result[0]
                    profile_data['last_login'] = result[1]
                    profile_data['is_active'] = result[2]
        else:
            with dbh.dbhLock:
                dbh.dbh.execute("""
                    SELECT mobile, created_at, last_login, is_active 
                    FROM tbl_users WHERE id = ?
                """, (user_info['id'],))
                result = dbh.dbh.fetchone()
                if result:
                    profile_data['mobile'] = result[0] or 'N/A'
                    profile_data['created_at'] = result[1]
                    profile_data['last_login'] = result[2]
                    profile_data['is_active'] = result[3]
        
        # Format dates
        profile_data['created_at_formatted'] = time.strftime("%B %d, %Y at %I:%M %p", time.localtime(profile_data.get('created_at', time.time())))
        profile_data['last_login_formatted'] = time.strftime("%B %d, %Y at %I:%M %p", time.localtime(profile_data.get('last_login', time.time()))) if profile_data.get('last_login') else 'Never'
        
        # Get statistics - only for this user/admin
        stats = {
            'total_scans': 0,
            'active_scans': 0,
            'total_activities': 0
        }
        
        with dbh.dbhLock:
            try:
                if user_info['type'] == 'admin':
                    # Get scans created by this admin
                    dbh.dbh.execute("""
                        SELECT COUNT(*) FROM tbl_scan_instance 
                        WHERE created_by_admin_id = ?
                    """, (user_info['id'],))
                    result = dbh.dbh.fetchone()
                    stats['total_scans'] = result[0] if result else 0
                    
                    dbh.dbh.execute("""
                        SELECT COUNT(*) FROM tbl_scan_instance 
                        WHERE created_by_admin_id = ? AND scan_status IN ('RUNNING', 'STARTING', 'STARTED')
                    """, (user_info['id'],))
                    result = dbh.dbh.fetchone()
                    stats['active_scans'] = result[0] if result else 0
                else:
                    # Get scans created by this user
                    dbh.dbh.execute("""
                        SELECT COUNT(*) FROM tbl_scan_instance 
                        WHERE created_by_user_id = ?
                    """, (user_info['id'],))
                    result = dbh.dbh.fetchone()
                    stats['total_scans'] = result[0] if result else 0
                    
                    dbh.dbh.execute("""
                        SELECT COUNT(*) FROM tbl_scan_instance 
                        WHERE created_by_user_id = ? AND scan_status IN ('RUNNING', 'STARTING', 'STARTED')
                    """, (user_info['id'],))
                    result = dbh.dbh.fetchone()
                    stats['active_scans'] = result[0] if result else 0
            except Exception:
                pass
        
        # Get recent activities
        activities = []
        activity_data = []
        
        with dbh.dbhLock:
            try:
                if user_info['type'] == 'admin':
                    dbh.dbh.execute("""
                        SELECT activity_type, activity_description, scan_id, created_at, ip_address 
                        FROM tbl_user_activity 
                        WHERE admin_id = ? 
                        ORDER BY created_at DESC
                    """, (user_info['id'],))
                else:
                    dbh.dbh.execute("""
                        SELECT activity_type, activity_description, scan_id, created_at, ip_address 
                        FROM tbl_user_activity 
                        WHERE user_id = ? 
                        ORDER BY created_at DESC
                    """, (user_info['id'],))
                activity_data = dbh.dbh.fetchall()
            except Exception:
                activity_data = []
        
        stats['total_activities'] = len(activity_data)
        
        for activity in activity_data[:10]:
            activities.append({
                'type': activity[0],
                'description': activity[1],
                'time_formatted': time.strftime("%B %d, %Y at %I:%M %p", time.localtime(activity[3]))
            })
        
        # Get session information
        session_id = cherrypy.session.get('session_id')
        session_info = {
            'ip_address': self.auth.get_client_ip(),
            'user_agent': self.auth.get_user_agent(),
            'created_at_formatted': 'Current session',
            'expires_at_formatted': 'Active'
        }
        
        if session_id:
            session_data = dbh.getSession(session_id)
            if session_data:
                session_info['created_at_formatted'] = time.strftime("%B %d, %Y at %I:%M %p", time.localtime(session_data['created_at']))
                session_info['expires_at_formatted'] = time.strftime("%B %d, %Y at %I:%M %p", time.localtime(session_data['expires_at']))
        
        templ = self.lookup.get_template('profile.tmpl')
        return templ.render(
            docroot=self.docroot,
            version=__version__,
            pageid='PROFILE',
            user_info=user_info,
            profile_data=profile_data,
            stats=stats,
            activities=activities,
            session_info=session_info
        )
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def update_profile(self, username: str, email: str, mobile: str = None) -> dict:
        """Update user/admin profile.
        
        Args:
            username (str): new username
            email (str): new email
            mobile (str): new mobile number
            
        Returns:
            dict: JSON response
        """
        user_info = self.auth.get_current_user()
        if not user_info:
            return {'success': False, 'message': 'Authentication required'}
        
        if not username or not email:
            return {'success': False, 'message': 'Username and email are required'}
        
        try:
            dbh = HawkEyeDb(self.config)
            
            with dbh.dbhLock:
                if user_info['type'] == 'admin':
                    dbh.dbh.execute("""
                        UPDATE tbl_admins 
                        SET username = ?, email = ?
                        WHERE id = ?
                    """, (username, email, user_info['id']))
                else:
                    dbh.dbh.execute("""
                        UPDATE tbl_users 
                        SET username = ?, email = ?, mobile = ?
                        WHERE id = ?
                    """, (username, email, mobile, user_info['id']))
                
                dbh.conn.commit()
            
            # Log the change
            dbh.logSystemEvent(
                'INFO',
                'USER_MANAGEMENT',
                f'{user_info["type"].capitalize()} "{user_info["username"]}" updated their profile',
                user_id=user_info['id'] if user_info['type'] == 'user' else None,
                admin_id=user_info['id'] if user_info['type'] == 'admin' else None,
                ip_address=self.auth.get_client_ip()
            )
            
            return {'success': True, 'message': 'Profile updated successfully'}
        except Exception as e:
            return {'success': False, 'message': f'Error updating profile: {str(e)}'}
    
    @cherrypy.expose
    @cherrypy.tools.json_out()
    def change_password(self, current_password: str, new_password: str) -> dict:
        """Change user/admin password.
        
        Args:
            current_password (str): current password
            new_password (str): new password
            
        Returns:
            dict: JSON response
        """
        user_info = self.auth.get_current_user()
        if not user_info:
            return {'success': False, 'message': 'Authentication required'}
        
        if not current_password or not new_password:
            return {'success': False, 'message': 'All fields are required'}
        
        if len(new_password) < 6:
            return {'success': False, 'message': 'Password must be at least 6 characters long'}
        
        try:
            dbh = HawkEyeDb(self.config)
            
            # Verify current password
            if user_info['type'] == 'admin':
                auth_result = dbh.authenticateAdmin(user_info['username'], current_password)
            else:
                auth_result = dbh.authenticateUser(user_info['username'], current_password)
            
            if not auth_result:
                return {'success': False, 'message': 'Current password is incorrect'}
            
            # Update password
            import hashlib
            password_hash = hashlib.sha256(new_password.encode()).hexdigest()
            
            with dbh.dbhLock:
                if user_info['type'] == 'admin':
                    dbh.dbh.execute("""
                        UPDATE tbl_admins 
                        SET password_hash = ?
                        WHERE id = ?
                    """, (password_hash, user_info['id']))
                else:
                    dbh.dbh.execute("""
                        UPDATE tbl_users 
                        SET password_hash = ?
                        WHERE id = ?
                    """, (password_hash, user_info['id']))
                
                dbh.conn.commit()
            
            # Log the change
            dbh.logSystemEvent(
                'WARNING',
                'AUTH',
                f'{user_info["type"].capitalize()} "{user_info["username"]}" changed their password',
                user_id=user_info['id'] if user_info['type'] == 'user' else None,
                admin_id=user_info['id'] if user_info['type'] == 'admin' else None,
                ip_address=self.auth.get_client_ip()
            )
            
            return {'success': True, 'message': 'Password changed successfully'}
        except Exception as e:
            return {'success': False, 'message': f'Error changing password: {str(e)}'}
    
    @cherrypy.expose
    def admin_export_system_logs(self, log_level: str = None, log_category: str = None) -> str:
        """Export system logs to CSV (admin only).
        
        Args:
            log_level (str): filter by log level
            log_category (str): filter by category
            
        Returns:
            str: CSV export of system logs
        """
        if not self.auth.require_auth('admin'):
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login/admin")
        
        try:
            dbh = HawkEyeDb(self.config)
            result = dbh.getSystemLogs(limit=10000, offset=0, log_level=log_level, log_category=log_category)
            
            csv_content = "Timestamp,Level,Category,Message,User,Admin,IP Address\n"
            for log in result['logs']:
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(log[9]))
                user = log[10] if log[10] else ''
                admin = log[11] if log[11] else ''
                ip = log[6] if log[6] else ''
                message = log[3].replace('"', '""')  # Escape quotes for CSV
                csv_content += f'"{timestamp}","{log[1]}","{log[2]}","{message}","{user}","{admin}","{ip}"\n'
            
            cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="hawkeye_system_logs.csv"'
            cherrypy.response.headers['Content-Type'] = 'text/csv'
            return csv_content
        except Exception as e:
            return f"Error exporting logs: {str(e)}"

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'HawkEyeWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = HawkEyeDb(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=HawkEye-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'HawkEyeWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = HawkEyeDb(self.config)

        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve info for scan."]).encode('utf-8')

        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve correlations for scan."]).encode('utf-8')

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-HawkEye-correlations.xlxs"
            else:
                fname = "HawkEye-correlations.xlxs"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, headings, sheetNameIndex=0)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)

            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-HawkEye-correlations.csv"
            else:
                fname = "HawkEye-correlations.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'HawkEyeWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = HawkEyeDb(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "HawkEye.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "HawkEye.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'HawkEyeWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = HawkEyeDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "HawkEye.xlsx"
            else:
                fname = scan_name + "-HawkEye.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "HawkEye.csv"
            else:
                fname = scan_name + "-HawkEye.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'HawkEyeWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=HawkEye.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=HawkEye.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'HawkEyeWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = HawkEyeDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<heURL>", "").replace("</heURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "HawkEye.json"
        else:
            fname = scan_name + "-HawkEye.json"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'HawkEyeWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = HawkEyeDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return HawkEyeHelpers.buildGraphJson([root], data)

        if not scan_name:
            fname = "HawkEye.gexf"
        else:
            fname = scan_name + "HawkEye.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return HawkEyeHelpers.buildGraphGexf([root], "HawkEye Export", data)

    @cherrypy.expose
    def scanvizmulti(self: 'HawkEyeWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = HawkEyeDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "HawkEye.gexf"
        else:
            fname = scan_name + "-HawkEye.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return HawkEyeHelpers.buildGraphGexf(roots, "HawkEye Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'HawkEyeWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = HawkEyeDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'HawkEyeWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = HawkEyeDb(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "hep__stor_stdout" in modlist:
            modlist.remove("hep__stor_stdout")

        targetType = HawkEyeHelpers.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = HawkEyeHelpers.targetTypeFromString(f'"{scantarget}"')

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = HawkEyeHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startHawkEyeScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'HawkEyeWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = HawkEyeDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "hep__stor_stdout" in modlist:
                modlist.remove("hep__stor_stdout")

            targetType = HawkEyeHelpers.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                return self.error("Invalid target type. Could not recognize it as a target HawkEye supports.")

            # Start running a new scan
            scanId = HawkEyeHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startHawkEyeScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] Scan [{scanId}] failed: {e}")
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = lookup.get_template('scanlist.tmpl')
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'HawkEyeWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        # Check if user is authenticated
        if not self.auth.require_auth():
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
        
        # Log user activity
        user_info = self.auth.get_current_user()
        if user_info:
            dbh = HawkEyeDb(self.config)
            dbh.logUserActivity(
                user_id=user_info['id'],
                activity_type='page_access',
                activity_description='Accessed new scan page',
                ip_address=self.auth.get_client_ip()
            )
        
        dbh = HawkEyeDb(self.config)
        types = dbh.eventTypes()
        templ = self.lookup.get_template('newscan.tmpl')
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__, user_info=user_info)

    @cherrypy.expose
    def clonescan(self: 'HawkEyeWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = HawkEyeDb(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = HawkEyeHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')
        
        # Get current user info
        user_info = self.auth.get_current_user()

        templ = self.lookup.get_template('newscan.tmpl')
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__, user_info=user_info)

    @cherrypy.expose
    def index(self: 'HawkEyeWebUi') -> str:
        """Show scan list page.
        
        Returns:
            str: Scan list page HTML
        """
        # Check if user is authenticated
        if not self.auth.require_auth():
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
        
        # Log user activity
        user_info = self.auth.get_current_user()
        if user_info:
            dbh = HawkEyeDb(self.config)
            dbh.logUserActivity(
                user_id=user_info['id'],
                activity_type='page_access',
                activity_description='Accessed scans page',
                ip_address=self.auth.get_client_ip()
            )
        
        templ = self.lookup.get_template('scanlist.tmpl')
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__, user_info=user_info)

    @cherrypy.expose
    def scaninfo(self: 'HawkEyeWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        # Get current user info
        user_info = self.auth.get_current_user()
        
        dbh = HawkEyeDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = self.lookup.get_template('scaninfo.tmpl')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST", user_info=user_info)

    @cherrypy.expose
    def opts(self: 'HawkEyeWebUi', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated succesheully

        Returns:
            str: scan options page HTML
        """
        # Get current user info
        user_info = self.auth.get_current_user()
        
        templ = self.lookup.get_template('opts.tmpl')
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot, user_info=user_info)

    @cherrypy.expose
    def moduleinfo(self: 'HawkEyeWebUi') -> str:
        """Show module information page.

        Returns:
            str: module information page HTML
        """
        # Check if user is authenticated
        if not self.auth.require_auth():
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
        
        # Get current user info
        user_info = self.auth.get_current_user()
        
        # Module data: [name, category, description, requiresAPIKey]
        module_data = json.dumps([
            ['helib', 'Core', 'Core HawkEye library functions', 'No'],
            ['hescan', 'Core', 'Main scanning engine', 'No'],
            ['hep_abstractapi', 'API Services', 'Email validation and IP geolocation via AbstractAPI', 'Yes'],
            ['hep_abusech', 'Threat Intelligence', 'Check against abuse.ch malware and botnet database', 'No'],
            ['hep_abuseipdb', 'Threat Intelligence', 'Check IP addresses against AbuseIPDB for malicious activity reports', 'Yes'],
            ['hep_abusix', 'Threat Intelligence', 'Query Abusix Mail Intelligence for email reputation', 'Yes'],
            ['hep_accounts', 'Social Media', 'Search for usernames across multiple social media platforms', 'No'],
            ['hep_adblock', 'Security', 'Check domains against adblock lists', 'No'],
            ['hep_adguard_dns', 'DNS', 'Query AdGuard DNS for domain filtering information', 'No'],
            ['hep_ahmia', 'Dark Web', 'Search Ahmia.fi for .onion sites', 'No'],
            ['hep_alienvault', 'Threat Intelligence', 'Query AlienVault OTX for threat intelligence data', 'Yes'],
            ['hep_alienvaultiprep', 'Reputation', 'Check IP reputation via AlienVault', 'No'],
            ['hep_apple_itunes', 'Search', 'Search Apple iTunes for apps and content', 'No'],
            ['hep_archiveorg', 'Web Archive', 'Search Archive.org Wayback Machine for historical website data', 'No'],
            ['hep_arin', 'Network', 'Query ARIN WHOIS for IP address registration information', 'No'],
            ['hep_azureblobstorage', 'Cloud Storage', 'Discover Azure Blob Storage containers', 'No'],
            ['hep_base64', 'Data Processing', 'Decode Base64 encoded strings', 'No'],
            ['hep_bgpview', 'Network', 'Query BGPView for BGP routing and ASN information', 'No'],
            ['hep_binaryedge', 'Search Engine', 'Search BinaryEdge for internet-wide scan data', 'Yes'],
            ['hep_bingsearch', 'Search Engine', 'Perform Bing web searches', 'Yes'],
            ['hep_bingsharedip', 'Network', 'Find websites sharing the same IP address via Bing', 'No'],
            ['hep_binstring', 'Data Processing', 'Extract binary strings from data', 'No'],
            ['hep_bitcoin', 'Cryptocurrency', 'Look up Bitcoin addresses and transactions', 'No'],
            ['hep_bitcoinabuse', 'Cryptocurrency', 'Check Bitcoin addresses against abuse database', 'No'],
            ['hep_bitcoinwhoswho', 'Cryptocurrency', 'Identify Bitcoin address owners', 'No'],
            ['hep_blockchain', 'Cryptocurrency', 'Query blockchain.com for Bitcoin data', 'No'],
            ['hep_blocklistde', 'Threat Intelligence', 'Check IPs against blocklist.de', 'No'],
            ['hep_botscout', 'Security', 'Check emails and IPs against BotScout spam database', 'Yes'],
            ['hep_botvrij', 'Threat Intelligence', 'Check against Botvrij.eu botnet database', 'No'],
            ['hep_builtwith', 'Web Technology', 'Identify web technologies used by websites', 'Yes'],
            ['hep_c99', 'API Services', 'Multi-purpose API for domain, IP, and phone lookups', 'Yes'],
            ['hep_callername', 'Phone', 'Look up phone number caller information', 'No'],
            ['hep_censys', 'Search Engine', 'Search Censys for internet-wide scan data', 'Yes'],
            ['hep_certspotter', 'SSL/TLS', 'Monitor SSL/TLS certificates via Certificate Transparency logs', 'Yes'],
            ['hep_cinsscore', 'Threat Intelligence', 'Check IPs against CINS Army threat list', 'No'],
            ['hep_circllu', 'Threat Intelligence', 'Query CIRCL.LU passive DNS and SSL databases', 'Yes'],
            ['hep_citadel', 'Threat Intelligence', 'Check against Citadel Trojan C&C servers', 'No'],
            ['hep_cleanbrowsing', 'DNS', 'Query CleanBrowsing DNS for filtering information', 'No'],
            ['hep_cleantalk', 'Security', 'Check emails and IPs against CleanTalk spam database', 'Yes'],
            ['hep_clearbit', 'Business Intelligence', 'Enrich company and person data via Clearbit', 'Yes'],
            ['hep_cloudflaredns', 'DNS', 'Query Cloudflare DNS', 'No'],
            ['hep_coinblocker', 'Security', 'Check domains against cryptocurrency mining blocklists', 'No'],
            ['hep_commoncrawl', 'Web Archive', 'Search Common Crawl web archive', 'No'],
            ['hep_comodo', 'SSL/TLS', 'Query Comodo certificate database', 'No'],
            ['hep_company', 'Business Intelligence', 'Extract company information from data', 'No'],
            ['hep_cookie', 'Web', 'Extract and analyze HTTP cookies', 'No'],
            ['hep_countryname', 'Geolocation', 'Identify country from various data types', 'No'],
            ['hep_creditcard', 'Data Processing', 'Identify credit card numbers in data', 'No'],
            ['hep_crobat_api', 'DNS', 'Query Project Sonar Crobat API for DNS data', 'No'],
            ['hep_crossref', 'Research', 'Search academic publications via Crossref', 'No'],
            ['hep_crt', 'SSL/TLS', 'Search crt.sh Certificate Transparency logs', 'No'],
            ['hep_crxcavator', 'Browser Extensions', 'Analyze Chrome extensions via CRXcavator', 'Yes'],
            ['hep_customfeed', 'Custom', 'Query custom threat intelligence feeds', 'No'],
            ['hep_cybercrimetracker', 'Threat Intelligence', 'Check against cybercrime-tracker.net C&C database', 'No'],
            ['hep_debounce', 'Email', 'Validate email addresses via Debounce', 'Yes'],
            ['hep_dehashed', 'Data Breach', 'Search Dehashed for breached credentials', 'Yes'],
            ['hep_digitaloceanspace', 'Cloud Storage', 'Discover DigitalOcean Spaces', 'No'],
            ['hep_dns_for_family', 'DNS', 'Query DNS for Family filtering service', 'No'],
            ['hep_dnsbrute', 'DNS', 'Brute force DNS subdomains', 'No'],
            ['hep_dnscommonsrv', 'DNS', 'Check for common DNS SRV records', 'No'],
            ['hep_dnsdb', 'DNS', 'Query Farsight DNSDB for passive DNS data', 'Yes'],
            ['hep_dnsdumpster', 'DNS', 'Search DNSdumpster for DNS records', 'No'],
            ['hep_dnsgrep', 'DNS', 'Search DNS records via DNSGrep', 'No'],
            ['hep_dnsneighbor', 'DNS', 'Find DNS neighbors (domains on same IP)', 'No'],
            ['hep_dnsraw', 'DNS', 'Perform raw DNS queries', 'No'],
            ['hep_dnsresolve', 'DNS', 'Resolve DNS records', 'No'],
            ['hep_dnszonexfer', 'DNS', 'Attempt DNS zone transfers', 'No'],
            ['hep_dronebl', 'Threat Intelligence', 'Check IPs against DroneBL', 'No'],
            ['hep_duckduckgo', 'Search Engine', 'Perform DuckDuckGo searches', 'No'],
            ['hep_email', 'Email', 'Extract email addresses from data', 'No'],
            ['hep_emailcrawlr', 'Email', 'Find email addresses via EmailCrawlr', 'Yes'],
            ['hep_emailformat', 'Email', 'Discover email format patterns for companies', 'No'],
            ['hep_emailrep', 'Email', 'Check email reputation via EmailRep', 'No'],
            ['hep_emergingthreats', 'Threat Intelligence', 'Check against Emerging Threats rules', 'No'],
            ['hep_errors', 'Data Processing', 'Extract error messages from data', 'No'],
            ['hep_ethereum', 'Cryptocurrency', 'Look up Ethereum addresses', 'No'],
            ['hep_etherscan', 'Cryptocurrency', 'Query Etherscan for Ethereum blockchain data', 'Yes'],
            ['hep_filemeta', 'File Analysis', 'Extract metadata from files', 'No'],
            ['hep_flickr', 'Social Media', 'Search Flickr for photos and users', 'Yes'],
            ['hep_focsec', 'Threat Intelligence', 'Query FOC Security threat feeds', 'No'],
            ['hep_fortinet', 'Threat Intelligence', 'Check against Fortinet threat intelligence', 'No'],
            ['hep_fraudguard', 'Fraud Detection', 'Check IPs and emails via FraudGuard', 'Yes'],
            ['hep_fsecure_riddler', 'DNS', 'Query F-Secure Riddler passive DNS', 'No'],
            ['hep_fullcontact', 'Contact Enrichment', 'Enrich contact information via FullContact', 'Yes'],
            ['hep_fullhunt', 'Attack Surface', 'Discover attack surface via FullHunt', 'Yes'],
            ['hep_github', 'Code Repository', 'Search GitHub for code, users, and repositories', 'Yes'],
            ['hep_gleif', 'Business Intelligence', 'Look up Legal Entity Identifiers (LEI)', 'No'],
            ['hep_google_tag_manager', 'Web Analytics', 'Identify Google Tag Manager usage', 'No'],
            ['hep_googlemaps', 'Geolocation', 'Query Google Maps for location data', 'Yes'],
            ['hep_googleobjectstorage', 'Cloud Storage', 'Discover Google Cloud Storage buckets', 'No'],
            ['hep_googlesafebrowsing', 'Security', 'Check URLs against Google Safe Browsing', 'Yes'],
            ['hep_googlesearch', 'Search Engine', 'Perform Google searches', 'Yes'],
            ['hep_gravatar', 'Social Media', 'Look up Gravatar profiles', 'No'],
            ['hep_grayhatwarfare', 'Cloud Storage', 'Search for exposed S3 buckets via GrayhatWarfare', 'Yes'],
            ['hep_greensnow', 'Threat Intelligence', 'Check IPs against GreenSnow blocklist', 'No'],
            ['hep_grep_app', 'Code Search', 'Search code repositories via grep.app', 'No'],
            ['hep_greynoise', 'Threat Intelligence', 'Query GreyNoise for internet scanner data', 'Yes'],
            ['hep_greynoise_community', 'Threat Intelligence', 'Query GreyNoise Community API', 'Yes'],
            ['hep_h1nobbdde', 'Threat Intelligence', 'Check against H1 NobbDe threat list', 'No'],
            ['hep_hackertarget', 'Network Tools', 'Use HackerTarget API for various network queries', 'Yes'],
            ['hep_hashes', 'Data Processing', 'Extract hash values from data', 'No'],
            ['hep_haveibeenpwned', 'Data Breach', 'Check emails against Have I Been Pwned', 'Yes'],
            ['hep_honeypot', 'Threat Intelligence', 'Identify honeypot systems', 'No'],
            ['hep_hosting', 'Network', 'Identify hosting providers', 'No'],
            ['hep_hostio', 'DNS', 'Query Host.io for domain and DNS data', 'Yes'],
            ['hep_hunter', 'Email', 'Find email addresses via Hunter.io', 'Yes'],
            ['hep_hybrid_analysis', 'Malware Analysis', 'Query Hybrid Analysis for malware reports', 'Yes'],
            ['hep_iban', 'Financial', 'Validate IBAN numbers', 'No'],
            ['hep_iknowwhatyoudownload', 'Torrent', 'Check what torrents an IP has downloaded', 'Yes'],
            ['hep_intelx', 'Search Engine', 'Search Intelligence X for leaked data', 'Yes'],
            ['hep_intfiles', 'Data Processing', 'Identify interesting files', 'No'],
            ['hep_ipapico', 'Geolocation', 'IP geolocation via ipapi.co', 'No'],
            ['hep_ipapicom', 'Geolocation', 'IP geolocation via ipapi.com', 'Yes'],
            ['hep_ipinfo', 'Geolocation', 'IP geolocation and ASN info via ipinfo.io', 'Yes'],
            ['hep_ipqualityscore', 'Fraud Detection', 'Check IP quality and fraud score', 'Yes'],
            ['hep_ipregistry', 'Geolocation', 'IP geolocation via IPRegistry', 'Yes'],
            ['hep_ipstack', 'Geolocation', 'IP geolocation via ipstack', 'Yes'],
            ['hep_isc', 'Threat Intelligence', 'Query ISC SANS Internet Storm Center', 'No'],
            ['hep_jsonwhoiscom', 'WHOIS', 'WHOIS lookups via jsonwhois.com', 'Yes'],
            ['hep_junkfiles', 'Web', 'Identify junk and backup files on web servers', 'No'],
            ['hep_keybase', 'Social Media', 'Look up Keybase profiles', 'No'],
            ['hep_koodous', 'Mobile Security', 'Query Koodous for Android app analysis', 'Yes'],
            ['hep_leakix', 'Data Leaks', 'Search LeakIX for exposed services and data', 'Yes'],
            ['hep_maltiverse', 'Threat Intelligence', 'Query Maltiverse threat intelligence', 'Yes'],
            ['hep_malwarepatrol', 'Threat Intelligence', 'Check against Malware Patrol lists', 'Yes'],
            ['hep_metadefender', 'Malware Analysis', 'Scan files and IPs via MetaDefender', 'Yes'],
            ['hep_mnemonic', 'Threat Intelligence', 'Query mnemonic passive DNS', 'Yes'],
            ['hep_multiproxy', 'Proxy Detection', 'Check if IP is a proxy', 'No'],
            ['hep_myspace', 'Social Media', 'Search MySpace profiles', 'No'],
            ['hep_nameapi', 'Name Analysis', 'Analyze person names via NameAPI', 'Yes'],
            ['hep_names', 'Data Processing', 'Extract person names from data', 'No'],
            ['hep_networksdb', 'Network', 'Query NetworksDB for network information', 'Yes'],
            ['hep_neutrinoapi', 'API Services', 'Multi-purpose API for various lookups', 'Yes'],
            ['hep_numverify', 'Phone', 'Validate phone numbers via Numverify', 'Yes'],
            ['hep_onioncity', 'Dark Web', 'Search Onion.City for .onion sites', 'No'],
            ['hep_onionsearchengine', 'Dark Web', 'Search onion search engines', 'No'],
            ['hep_onyphe', 'Search Engine', 'Query Onyphe for internet-wide scan data', 'Yes'],
            ['hep_openbugbounty', 'Security', 'Check against Open Bug Bounty database', 'No'],
            ['hep_opencorporates', 'Business Intelligence', 'Search OpenCorporates for company data', 'Yes'],
            ['hep_opendns', 'DNS', 'Query OpenDNS', 'No'],
            ['hep_opennic', 'DNS', 'Query OpenNIC DNS', 'No'],
            ['hep_openphish', 'Threat Intelligence', 'Check against OpenPhish phishing database', 'No'],
            ['hep_openstreetmap', 'Geolocation', 'Query OpenStreetMap for location data', 'No'],
            ['hep_pageinfo', 'Web', 'Extract page information from websites', 'No'],
            ['hep_pastebin', 'Data Leaks', 'Search Pastebin for leaked data', 'Yes'],
            ['hep_pgp', 'Encryption', 'Search PGP key servers', 'No'],
            ['hep_phishstats', 'Threat Intelligence', 'Check against PhishStats phishing database', 'No'],
            ['hep_phishtank', 'Threat Intelligence', 'Check against PhishTank phishing database', 'Yes'],
            ['hep_phone', 'Phone', 'Extract phone numbers from data', 'No'],
            ['hep_portscan_tcp', 'Network', 'Perform TCP port scans', 'No'],
            ['hep_projectdiscovery', 'DNS', 'Query Project Discovery Chaos for DNS data', 'Yes'],
            ['hep_psbdmp', 'Data Leaks', 'Search psbdmp for Pastebin dumps', 'No'],
            ['hep_pulsedive', 'Threat Intelligence', 'Query Pulsedive threat intelligence', 'Yes'],
            ['hep_punkspider', 'Security', 'Search PunkSpider for website vulnerabilities', 'No'],
            ['hep_quad9', 'DNS', 'Query Quad9 DNS', 'No'],
            ['hep_reversewhois', 'WHOIS', 'Perform reverse WHOIS lookups', 'No'],
            ['hep_ripe', 'Network', 'Query RIPE NCC for IP and ASN information', 'No'],
            ['hep_riskiq', 'Threat Intelligence', 'Query RiskIQ for threat data', 'Yes'],
            ['hep_robtex', 'Network', 'Query Robtex for DNS and network data', 'No'],
            ['hep_s3bucket', 'Cloud Storage', 'Discover Amazon S3 buckets', 'No'],
            ['hep_searchcode', 'Code Search', 'Search source code via searchcode', 'No'],
            ['hep_securitytrails', 'DNS', 'Query SecurityTrails for DNS history', 'Yes'],
            ['hep_seon', 'Fraud Detection', 'Check emails and phones via SEON', 'Yes'],
            ['hep_shodan', 'Search Engine', 'Query Shodan for internet-connected devices', 'Yes'],
            ['hep_similar', 'Web', 'Find similar websites', 'No'],
            ['hep_skymem', 'Email', 'Search Skymem for email addresses', 'No'],
            ['hep_slideshare', 'Document Search', 'Search SlideShare for presentations', 'No'],
            ['hep_snov', 'Email', 'Find email addresses via Snov.io', 'Yes'],
            ['hep_social', 'Social Media', 'Extract social media links', 'No'],
            ['hep_sociallinks', 'Social Media', 'Find social media profiles via SocialLinks', 'Yes'],
            ['hep_socialprofiles', 'Social Media', 'Search for social media profiles', 'No'],
            ['hep_sorbs', 'Threat Intelligence', 'Check IPs against SORBS blocklist', 'No'],
            ['hep_spamcop', 'Threat Intelligence', 'Check IPs against SpamCop blocklist', 'No'],
            ['hep_spamhaus', 'Threat Intelligence', 'Check IPs against Spamhaus blocklists', 'No'],
            ['hep_spider', 'Web Crawling', 'Spider websites for links and content', 'No'],
            ['hep_spur', 'Network', 'Query Spur for IP intelligence', 'Yes'],
            ['hep_spyonweb', 'Web Intelligence', 'Find related websites via SpyOnWeb', 'Yes'],
            ['hep_sslcert', 'SSL/TLS', 'Extract SSL/TLS certificate information', 'No'],
            ['hep_stackoverflow', 'Developer', 'Search Stack Overflow', 'No'],
            ['hep_stevenblack_hosts', 'Security', 'Check against Steven Black hosts file', 'No'],
            ['hep_strangeheaders', 'Web', 'Identify unusual HTTP headers', 'No'],
            ['hep_subdomain_takeover', 'Security', 'Check for subdomain takeover vulnerabilities', 'No'],
            ['hep_sublist3r', 'DNS', 'Enumerate subdomains via Sublist3r', 'No'],
            ['hep_surbl', 'Threat Intelligence', 'Check domains against SURBL', 'No'],
            ['hep_talosintel', 'Threat Intelligence', 'Query Cisco Talos Intelligence', 'No'],
            ['hep_textmagic', 'Phone', 'Validate phone numbers via TextMagic', 'Yes'],
            ['hep_threatcrowd', 'Threat Intelligence', 'Query ThreatCrowd for threat data', 'No'],
            ['hep_threatfox', 'Threat Intelligence', 'Check against abuse.ch ThreatFox', 'No'],
            ['hep_threatjammer', 'Threat Intelligence', 'Query ThreatJammer threat feeds', 'Yes'],
            ['hep_threatminer', 'Threat Intelligence', 'Query ThreatMiner for threat data', 'No'],
            ['hep_tldsearch', 'DNS', 'Search across TLDs for domain variations', 'No'],
            ['hep_tool_cmseek', 'Web Security', 'Detect CMS and vulnerabilities via CMSeeK', 'No'],
            ['hep_tool_dnstwist', 'DNS', 'Find typosquatting domains via dnstwist', 'No'],
            ['hep_tool_nbtscan', 'Network', 'Scan for NetBIOS information', 'No'],
            ['hep_tool_nmap', 'Network', 'Perform Nmap port scans', 'No'],
            ['hep_tool_nuclei', 'Web Security', 'Scan for vulnerabilities via Nuclei', 'No'],
            ['hep_tool_onesixtyone', 'Network', 'Scan for SNMP information', 'No'],
            ['hep_tool_retirejs', 'Web Security', 'Detect vulnerable JavaScript libraries', 'No'],
            ['hep_tool_snallygaster', 'Web Security', 'Find secret files on web servers', 'No'],
            ['hep_tool_testsslsh', 'SSL/TLS', 'Test SSL/TLS configuration via testssl.sh', 'No'],
            ['hep_tool_trufflehog', 'Security', 'Find secrets in Git repositories', 'No'],
            ['hep_tool_wafw00f', 'Web Security', 'Detect Web Application Firewalls', 'No'],
            ['hep_tool_wappalyzer', 'Web Technology', 'Identify web technologies via Wappalyzer', 'No'],
            ['hep_tool_whatweb', 'Web Technology', 'Identify web technologies via WhatWeb', 'No'],
            ['hep_torch', 'Dark Web', 'Search Torch for .onion sites', 'No'],
            ['hep_torexits', 'Network', 'Identify Tor exit nodes', 'No'],
            ['hep_trashpanda', 'Data Leaks', 'Search for exposed data via TrashPanda', 'No'],
            ['hep_trumail', 'Email', 'Validate email addresses via Trumail', 'No'],
            ['hep_twilio', 'Phone', 'Look up phone numbers via Twilio', 'Yes'],
            ['hep_twitter', 'Social Media', 'Search Twitter for users and tweets', 'Yes'],
            ['hep_uceprotect', 'Threat Intelligence', 'Check IPs against UCEPROTECT blocklists', 'No'],
            ['hep_urlscan', 'Web Security', 'Scan URLs via urlscan.io', 'Yes'],
            ['hep_venmo', 'Social Media', 'Search Venmo for payment transactions', 'No'],
            ['hep_viewdns', 'DNS', 'Various DNS tools via ViewDNS', 'Yes'],
            ['hep_virustotal', 'Malware Analysis', 'Scan files and URLs via VirusTotal', 'Yes'],
            ['hep_voipbl', 'Threat Intelligence', 'Check IPs against VoIP Blacklist', 'No'],
            ['hep_vxvault', 'Malware Analysis', 'Check against VX Vault malware database', 'No'],
            ['hep_webanalytics', 'Web Analytics', 'Identify web analytics tools', 'No'],
            ['hep_webframework', 'Web Technology', 'Identify web frameworks', 'No'],
            ['hep_webserver', 'Web Technology', 'Identify web servers', 'No'],
            ['hep_whatcms', 'Web Technology', 'Identify CMS via WhatCMS', 'Yes'],
            ['hep_whois', 'WHOIS', 'Perform WHOIS lookups', 'No'],
            ['hep_whoisology', 'WHOIS', 'Reverse WHOIS via Whoisology', 'Yes'],
            ['hep_whoxy', 'WHOIS', 'WHOIS and reverse WHOIS via Whoxy', 'Yes'],
            ['hep_wigle', 'Wireless', 'Search WiGLE for WiFi networks', 'Yes'],
            ['hep_wikileaks', 'Data Leaks', 'Search WikiLeaks', 'No'],
            ['hep_wikipediaedits', 'Research', 'Find Wikipedia edits by IP', 'No'],
            ['hep_xforce', 'Threat Intelligence', 'Query IBM X-Force Exchange', 'Yes'],
            ['hep_yandexdns', 'DNS', 'Query Yandex DNS', 'No'],
            ['hep_zetalytics', 'DNS', 'Query Zetalytics for DNS data', 'Yes'],
            ['hep_zonefiles', 'DNS', 'Search DNS zone files', 'Yes'],
            ['hep_zoneh', 'Security', 'Check against Zone-H defacement database', 'No']
        ])
        
        templ = self.lookup.get_template('moduleinfo.tmpl')
        return templ.render(pageid='MODULEINFO', version=__version__, docroot=self.docroot, user_info=user_info, module_data=module_data)

    @cherrypy.expose
    def optsexport(self: 'HawkEyeWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        he = HawkEye(self.config)
        conf = he.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="HawkEye.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'HawkEyeWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'HawkEyeWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = HawkEyeDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'HawkEyeWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from HawkEye? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = HawkEyeDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            he = HawkEye(self.config)
            self.config = he.configUnserialize(cleanopts, currentopts)
            dbh.configSet(he.configSerialize(self.config))
        except Exception as e:
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'HawkEyeWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = HawkEyeDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            he = HawkEye(self.config)
            self.config = he.configUnserialize(cleanopts, currentopts)
            dbh.configSet(he.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'HawkEyeWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = HawkEyeDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'HawkEyeWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = HawkEyeDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'HawkEyeWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = HawkEyeDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'HawkEyeWebUi') -> list:
        """List all modules.

        Returns:
            list: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret

        modinfo.sort()

        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'HawkEyeWebUi') -> list:
        """List all correlation rules.

        Returns:
            list: list of correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        rules = self.config['__correlationrules__']
        if not rules:
            return ret

        for r in rules:
            ret.append({
                'id': r['id'],
                'name': r['meta']['name'],
                'descr': r['meta']['description'],
                'risk': r['meta']['risk'],
            })

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'HawkEyeWebUi') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: HawkEye version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'HawkEyeWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = HawkEyeDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'HawkEyeWebUi', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        # Check if user is authenticated
        if not self.auth.require_auth():
            raise cherrypy.HTTPRedirect(f"{self.docroot}/login")
        
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan name was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan name was not specified.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: scan target was not specified."]).encode('utf-8')

            return self.error("Invalid request: scan target was not specified.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        targetType = HawkEyeHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Unrecognised target type."]).encode('utf-8')

            return self.error("Invalid target type. Could not recognize it as a target HawkEye supports.")

        # Swap the globalscantable for the database handler
        dbh = HawkEyeDb(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        he = HawkEye(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = he.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in he.eventsToModules(newmodcpy):
                    xmods = he.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "Incorrect usage: no modules specified for scan."]).encode('utf-8')

            return self.error("Invalid request: no modules specified for scan.")

        # Add our mandatory storage module
        if "hep__stor_db" not in modlist:
            modlist.append("hep__stor_db")
        modlist.sort()

        # Delete the stdout module in case it crept in
        if "hep__stor_stdout" in modlist:
            modlist.remove("hep__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = HawkEyeHelpers.genScanInstanceId()
        
        # Get current user info to track scan ownership
        user_info = self.auth.get_current_user()
        created_by_user_id = None
        created_by_admin_id = None
        
        if user_info:
            if user_info['type'] == 'user':
                created_by_user_id = user_info['id']
            elif user_info['type'] == 'admin':
                created_by_admin_id = user_info['id']
        
        try:
            p = mp.Process(target=startHawkEyeScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg, True, created_by_user_id, created_by_admin_id))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)
        
        # Log scan creation activity
        if user_info:
            dbh.logUserActivity(
                user_id=user_info['id'],
                activity_type='scan_created',
                activity_description=f'Created scan: {scanname} (target: {scantarget})',
                scan_id=scanId,
                ip_address=self.auth.get_client_ip()
            )

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'HawkEyeWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = HawkEyeDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart HawkEye.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        dbh = HawkEyeDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'HawkEyeWebUi', id: str, limit: str = "500", rowId: str = None, reverse: str = None, logType: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): limit number of results (default 500, "0" for all logs)
            rowId (str): retrieve logs starting from row ID
            reverse (str): search result order
            logType (str): filter by log type (INFO, WARNING, ERROR, DEBUG)

        Returns:
            list: scan log
        """
        dbh = HawkEyeDb(self.config)
        retdata = []

        try:
            # Convert limit to int
            # "0" or 0 means all logs (no limit)
            limit_int = int(limit) if limit else 500
            
            # If limit is 0, treat as None (all logs)
            if limit_int == 0:
                limit_int = None
            
            rowId_int = int(rowId) if rowId else 0
            reverse_bool = reverse == "1" if reverse else False
            
            data = dbh.scanLogs(id, limit_int, rowId_int, reverse_bool, logType)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'HawkEyeWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = HawkEyeDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'HawkEyeWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = HawkEyeDb(self.config)
        
        # Get current user info to filter scans
        user_info = self.auth.get_current_user()
        user_id = None
        admin_id = None
        
        if user_info:
            if user_info['type'] == 'user':
                user_id = user_info['id']
            elif user_info['type'] == 'admin':
                admin_id = user_info['id']
        
        # Filter scans by user/admin
        data = dbh.scanInstanceList(user_id=user_id, admin_id=admin_id)
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'HawkEyeWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = HawkEyeDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'HawkEyeWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = HawkEyeDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'HawkEyeWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list
        """
        retdata = []

        dbh = HawkEyeDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'HawkEyeWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results
        """
        retdata = []

        dbh = HawkEyeDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'HawkEyeWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = HawkEyeDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'HawkEyeWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'HawkEyeWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = HawkEyeDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'HawkEyeWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = HawkEyeDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = HawkEyeHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scangeodata(self: 'HawkEyeWebUi', id: str) -> dict:
        """Get geographical location data for a scan.

        Args:
            id (str): scan ID

        Returns:
            dict: geographical data with locations and coordinates
        """
        import re
        
        dbh = HawkEyeDb(self.config)
        retdata = {
            'locations': [],
            'unique_count': 0,
            'total_count': 0
        }

        try:
            # Get all GEOINFO events
            geoinfo_data = dbh.scanResultEvent(id, 'GEOINFO')
            
            # Dictionary to track unique locations and their counts
            location_map = {}
            
            for row in geoinfo_data:
                location = row[1]  # The data field contains location info
                
                if location and location not in location_map:
                    location_map[location] = {
                        'location': location,
                        'count': 1,
                        'coordinates': None,
                        'source': row[2] if len(row) > 2 else 'Unknown'
                    }
                elif location:
                    location_map[location]['count'] += 1
            
            # Try to get coordinates using a geocoding service or predefined mappings
            # For now, we'll use a simple approach with common locations
            # In production, you'd use a geocoding API like OpenCage, Nominatim, etc.
            
            for loc_key, loc_data in location_map.items():
                # Try to extract coordinates if available in the data
                # Some modules might provide coordinates
                coords = self._extract_coordinates(loc_data['location'])
                if coords:
                    loc_data['coordinates'] = coords
                else:
                    # Try to geocode the location
                    coords = self._geocode_location(loc_data['location'])
                    if coords:
                        loc_data['coordinates'] = coords
                
                retdata['locations'].append(loc_data)
            
            retdata['unique_count'] = len(location_map)
            retdata['total_count'] = sum(loc['count'] for loc in retdata['locations'])
            
        except Exception as e:
            self.log.error(f"Error fetching geo data: {e}")
            return retdata

        return retdata
    
    def _extract_coordinates(self, location_str: str) -> dict:
        """Extract coordinates from location string if present.
        
        Args:
            location_str (str): location string that might contain coordinates
            
        Returns:
            dict: {'lat': float, 'lon': float} or None
        """
        import re
        
        # Pattern to match coordinates like "40.7128, -74.0060" or "(40.7128, -74.0060)"
        coord_pattern = r'[-+]?\d*\.?\d+\s*,\s*[-+]?\d*\.?\d+'
        match = re.search(coord_pattern, location_str)
        
        if match:
            try:
                coords = match.group().split(',')
                lat = float(coords[0].strip())
                lon = float(coords[1].strip())
                
                # Validate coordinates
                if -90 <= lat <= 90 and -180 <= lon <= 180:
                    return {'lat': lat, 'lon': lon}
            except (ValueError, IndexError):
                pass
        
        return None
    
    def _geocode_location(self, location: str) -> dict:
        """Geocode a location string to coordinates using a simple mapping.
        
        Args:
            location (str): location string (e.g., "New York, US" or "London, GB")
            
        Returns:
            dict: {'lat': float, 'lon': float} or None
        """
        # Simple mapping of common locations
        # In production, use a proper geocoding API
        location_lower = location.lower()
        
        # Common country coordinates (capital cities)
        country_coords = {
            'united states': {'lat': 37.0902, 'lon': -95.7129},
            'us': {'lat': 37.0902, 'lon': -95.7129},
            'usa': {'lat': 37.0902, 'lon': -95.7129},
            'united kingdom': {'lat': 51.5074, 'lon': -0.1278},
            'uk': {'lat': 51.5074, 'lon': -0.1278},
            'gb': {'lat': 51.5074, 'lon': -0.1278},
            'canada': {'lat': 45.4215, 'lon': -75.6972},
            'ca': {'lat': 45.4215, 'lon': -75.6972},
            'germany': {'lat': 52.5200, 'lon': 13.4050},
            'de': {'lat': 52.5200, 'lon': 13.4050},
            'france': {'lat': 48.8566, 'lon': 2.3522},
            'fr': {'lat': 48.8566, 'lon': 2.3522},
            'india': {'lat': 28.6139, 'lon': 77.2090},
            'in': {'lat': 28.6139, 'lon': 77.2090},
            'china': {'lat': 39.9042, 'lon': 116.4074},
            'cn': {'lat': 39.9042, 'lon': 116.4074},
            'japan': {'lat': 35.6762, 'lon': 139.6503},
            'jp': {'lat': 35.6762, 'lon': 139.6503},
            'australia': {'lat': -35.2809, 'lon': 149.1300},
            'au': {'lat': -35.2809, 'lon': 149.1300},
            'brazil': {'lat': -15.8267, 'lon': -47.9218},
            'br': {'lat': -15.8267, 'lon': -47.9218},
            'russia': {'lat': 55.7558, 'lon': 37.6173},
            'ru': {'lat': 55.7558, 'lon': 37.6173},
            'netherlands': {'lat': 52.3676, 'lon': 4.9041},
            'nl': {'lat': 52.3676, 'lon': 4.9041},
            'spain': {'lat': 40.4168, 'lon': -3.7038},
            'es': {'lat': 40.4168, 'lon': -3.7038},
            'italy': {'lat': 41.9028, 'lon': 12.4964},
            'it': {'lat': 41.9028, 'lon': 12.4964},
            'mexico': {'lat': 19.4326, 'lon': -99.1332},
            'mx': {'lat': 19.4326, 'lon': -99.1332},
            'south korea': {'lat': 37.5665, 'lon': 126.9780},
            'kr': {'lat': 37.5665, 'lon': 126.9780},
            'singapore': {'lat': 1.3521, 'lon': 103.8198},
            'sg': {'lat': 1.3521, 'lon': 103.8198},
        }
        
        # City coordinates
        city_coords = {
            'new york': {'lat': 40.7128, 'lon': -74.0060},
            'london': {'lat': 51.5074, 'lon': -0.1278},
            'paris': {'lat': 48.8566, 'lon': 2.3522},
            'tokyo': {'lat': 35.6762, 'lon': 139.6503},
            'beijing': {'lat': 39.9042, 'lon': 116.4074},
            'moscow': {'lat': 55.7558, 'lon': 37.6173},
            'sydney': {'lat': -33.8688, 'lon': 151.2093},
            'berlin': {'lat': 52.5200, 'lon': 13.4050},
            'mumbai': {'lat': 19.0760, 'lon': 72.8777},
            'delhi': {'lat': 28.6139, 'lon': 77.2090},
            'shanghai': {'lat': 31.2304, 'lon': 121.4737},
            'los angeles': {'lat': 34.0522, 'lon': -118.2437},
            'chicago': {'lat': 41.8781, 'lon': -87.6298},
            'toronto': {'lat': 43.6532, 'lon': -79.3832},
            'amsterdam': {'lat': 52.3676, 'lon': 4.9041},
            'madrid': {'lat': 40.4168, 'lon': -3.7038},
            'rome': {'lat': 41.9028, 'lon': 12.4964},
            'singapore': {'lat': 1.3521, 'lon': 103.8198},
            'hong kong': {'lat': 22.3193, 'lon': 114.1694},
            'dubai': {'lat': 25.2048, 'lon': 55.2708},
            'san francisco': {'lat': 37.7749, 'lon': -122.4194},
            'seattle': {'lat': 47.6062, 'lon': -122.3321},
            'boston': {'lat': 42.3601, 'lon': -71.0589},
            'washington': {'lat': 38.9072, 'lon': -77.0369},
            'miami': {'lat': 25.7617, 'lon': -80.1918},
        }
        
        # Try to match city first
        for city, coords in city_coords.items():
            if city in location_lower:
                return coords
        
        # Try to match country
        for country, coords in country_coords.items():
            if country in location_lower:
                return coords
        
        return None
